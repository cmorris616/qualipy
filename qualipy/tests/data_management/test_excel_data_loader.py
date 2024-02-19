import datetime
import os
from unittest import TestCase

from openpyxl import Workbook

from qualipy.data_management.excel_data_loader import ExcelDataLoader


class TestExcelDataLoader(TestCase):
    def setUp(self) -> None:
        self.excel_file_name = 'unittest_excel_dl_data.xlsx'

        wb = Workbook()

        ws = wb.create_sheet(title='data_management.test_excel_data_loader.Employee')
        row = 1
        ws.cell(row=row, column=1, value='id')
        ws.cell(row=row, column=2, value='first_name')
        ws.cell(row=row, column=3, value='last_name')
        ws.cell(row=row, column=4, value='active')
        ws.cell(row=row, column=5, value='start_date')
        ws.cell(row=row, column=6, value='height')
        ws.cell(row=row, column=7, value='activation')
        ws.cell(row=row, column=8, value='shift_end')

        row = 2
        ws.cell(row=row, column=1, value=4)
        ws.cell(row=row, column=2, value='Bob')
        ws.cell(row=row, column=3, value='Jones')
        ws.cell(row=row, column=4, value=True)
        ws.cell(row=row, column=5, value=datetime.date(year=2019, month=2, day=5))
        ws.cell(row=row, column=6, value=5.5)
        ws.cell(row=row, column=7, value=datetime.datetime(year=2019, month=2, day=5, hour=8, minute=0, second=0))
        ws.cell(row=row, column=8, value=datetime.time(hour=17, minute=0, second=0))

        row = 3
        ws.cell(row=row, column=1, value=9)
        ws.cell(row=row, column=2, value='Alice')
        ws.cell(row=row, column=3, value='Baker')
        ws.cell(row=row, column=4, value=False)
        ws.cell(row=row, column=5, value=datetime.date(year=2020, month=12, day=5))
        ws.cell(row=row, column=6, value=5.25)
        ws.cell(row=row, column=7, value=datetime.datetime(year=2020, month=12, day=5, hour=9, minute=0, second=0))
        ws.cell(row=row, column=8, value=datetime.time(hour=18, minute=0, second=0))

        ws = wb.create_sheet(title='data_management.test_excel_data_loader.Task')

        row=1
        ws.cell(row=row, column=1, value='id')
        ws.cell(row=row, column=2, value='description')
        ws.cell(row=row, column=3, value='due_date')
        ws.cell(row=row, column=4, value='completion_datetime')
        ws.cell(row=row, column=5, value='start_date')
        ws.cell(row=row, column=6, value='start_time')

        row=2
        ws.cell(row=row, column=1, value=3)
        ws.cell(row=row, column=2, value='Add more data loaders')
        ws.cell(row=row, column=3, value=datetime.date(year=2024, month=10, day=30))
        ws.cell(row=row, column=4, value=datetime.datetime(year=2024, month=4, day=20, hour=3, minute=56, second=0))
        ws.cell(row=row, column=5, value=datetime.date(year=2024, month=1, day=15))
        ws.cell(row=row, column=6, value=datetime.time(hour=13, minute=20, second=0))

        wb.remove(wb.worksheets[0])

        wb.save(self.excel_file_name)

    def tearDown(self) -> None:
        os.remove(self.excel_file_name)

    def test_load_data(self):
        dl = ExcelDataLoader()
        
        records = dl.load_data(data_source=self.excel_file_name)
        assert isinstance(records, list)
        from data_management.test_excel_data_loader import Employee, Task
        assert len(records) == 3
        assert isinstance(records[0], Employee)
        assert isinstance(records[1], Employee)
        assert isinstance(records[2], Task)

        bob = records[0]
        alice = records[1]
        task = records[2]

        assert bob.id == 4
        assert bob.first_name == 'Bob'
        assert bob.last_name == 'Jones'
        assert bob.active
        assert bob.start_date == datetime.date(year=2019, month=2, day=5)
        assert bob.height == 5.5
        assert bob.activation == datetime.datetime(year=2019, month=2, day=5, hour=8, minute=0, second=0)
        assert bob.shift_end == datetime.time(hour=17, minute=0, second=0)

        assert alice.id == 9
        assert alice.first_name == 'Alice'
        assert alice.last_name == 'Baker'
        assert not alice.active
        assert alice.start_date == datetime.date(year=2020, month=12, day=5)
        assert alice.height == 5.25
        assert alice.activation == datetime.datetime(year=2020, month=12, day=5, hour=9, minute=0, second=0)
        assert alice.shift_end == datetime.time(hour=18, minute=0, second=0)

        assert task.id == 3
        assert task.description == 'Add more data loaders'
        assert task.due_date == datetime.date(year=2024, month=10, day=30)
        assert task.completion_datetime == datetime.datetime(year=2024, month=4, day=20, hour=3, minute=56, second=0)
        assert task.start_date == datetime.date(year=2024, month=1, day=15)
        assert task.start_time == datetime.time(hour=13, minute=20, second=0)

class Employee:
    def __init__(self) -> None:
        self.first_name: str = ""
        self.last_name: str = ""
        self.id: int = -1
        self.active: bool = False
        self.start_date: datetime.date = datetime.datetime.strptime('1970-01-01', '%Y-%m-%d').date()
        self.height: float = 0.0
        self.activation: datetime.datetime = datetime.datetime.strptime('1970-01-01 8:00:00', '%Y-%m-%d %I:%M:%S')
        self.shift_end: datetime.time = datetime.datetime.strptime('1970-01-01 8:00:00', '%Y-%m-%d %I:%M:%S').time()


class Task:
    def __init__(self) -> None:
        self.id: int = -1
        self.description: str = ""
        self.due_date: datetime.date = datetime.date.max
        self.completion_datetime: datetime.datetime = datetime.datetime.max
        self.start_time: datetime.time = datetime.time.max
        self.start_date: datetime.date = datetime.date.max